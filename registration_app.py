"""
registration_app.py
-------------------
MK30 Desktop App — two tabs:
  Tab 1: Registration Generator  — upload export → generate BOM + part list + registration CSV
  Tab 2: 44807 vs 44807          — compare our BOM's 44807 sheet against the website's export

Requirements:
    pip install customtkinter openpyxl pandas pdfplumber

Run:
    python registration_app.py
"""

import csv
import os
import re
import threading
from tkinter import filedialog, messagebox

import customtkinter as ctk
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

ANCHOR_PN = "LBL-F5-01"

CSV_HEADERS = [
    "Serial Number", "Tail Number", "Facility", "Model Number",
    "Block Number", "Software Version", "Knobs Mapping Key",
    "Min Camera Temp C", "Max Camera Temp C",
]
CSV_DEFAULTS = [
    "", "", "SEA109", "MK30", "",
    "DroneIronBirdHILSim-AdnCX3-block20-rl2at6364601745-MK30BL20-DevRel-8.2.B-signed-commercial",
    "I-1", "0", "40",
]

# Embedded serial map (pre-extracted from MK30_Serialization.pdf)
SERIAL_MAP = {"1786501001": {"tail": "N109PA", "block": ""}, "1786501002": {"tail": "N138PA", "block": ""}, "1786501003": {"tail": "N139PA", "block": ""}, "1786501004": {"tail": "N145PA", "block": ""}, "1786501005": {"tail": "N146PA", "block": ""}, "1786501006": {"tail": "N997PA", "block": ""}, "1786501007": {"tail": "N152PA", "block": ""}, "1786501008": {"tail": "N159PA", "block": ""}, "1786501009": {"tail": "N162PA", "block": ""}, "178650100A": {"tail": "N168PA", "block": ""}, "178650100B": {"tail": "N174PA", "block": ""}, "178650100C": {"tail": "N179PA", "block": ""}, "178650100D": {"tail": "N181PA", "block": ""}, "178650100E": {"tail": "N202PA", "block": "10"}, "178650100F": {"tail": "N206PA", "block": "10"}, "178650100G": {"tail": "N209PA", "block": "10"}, "178650100H": {"tail": "N213PA", "block": "10"}, "178650100J": {"tail": "N214PA", "block": "10"}, "178650100K": {"tail": "N219PA", "block": "10"}, "178650100L": {"tail": "N220PA", "block": "10"}, "178650100M": {"tail": "N223PA", "block": "10"}, "178650100N": {"tail": "N228PA", "block": "10"}, "176850100P": {"tail": "", "block": ""}, "178650100Q": {"tail": "N238PA", "block": "10"}, "178650100R": {"tail": "N254PA", "block": "10"}, "178650100S": {"tail": "N256PA", "block": "10"}, "178650100T": {"tail": "N257PA", "block": "10"}, "178650100U": {"tail": "N265PA", "block": "10"}, "178650100V": {"tail": "N274PA", "block": "10"}, "178650100W": {"tail": "N276PA", "block": "10"}, "178650100X": {"tail": "N282PA", "block": "10"}, "1786501010": {"tail": "N285PA", "block": "10"}, "1786501011": {"tail": "N293PA", "block": "10"}, "1786501012": {"tail": "N295PA", "block": "10"}, "1786501013": {"tail": "N322PA", "block": "10"}, "1786501014": {"tail": "N337PA", "block": "10"}, "1786501015": {"tail": "N339PA", "block": "10"}, "1786501016": {"tail": "N341PA", "block": "10"}, "1786501017": {"tail": "N386PA", "block": "10"}, "1786501018": {"tail": "N387PA", "block": "10"}, "1786501019": {"tail": "N398PA", "block": "10"}, "178650101A": {"tail": "N399PA", "block": "10"}, "178650101B": {"tail": "N471PA", "block": "10"}, "178650101C": {"tail": "N473PA", "block": "10"}, "178650101D": {"tail": "N503PA", "block": "10"}, "178650101E": {"tail": "N505PA", "block": "10"}, "178650101F": {"tail": "N543PA", "block": "10"}, "178650101G": {"tail": "N558PA", "block": "10"}, "178650101H": {"tail": "N559PA", "block": "10"}, "178650101J": {"tail": "N570PA", "block": "10"}, "178650101K": {"tail": "N579PA", "block": "10"}, "178650101L": {"tail": "N787PA", "block": "10"}, "178650101M": {"tail": "N791PA", "block": "10"}, "178650101N": {"tail": "N792PA", "block": "10"}, "178650101P": {"tail": "", "block": ""}, "178650101Q": {"tail": "N793PA", "block": "10"}, "178650101R": {"tail": "N796PA", "block": "10"}, "178650101S": {"tail": "N797PA", "block": "10"}, "178650101T": {"tail": "N831PA", "block": "10"}, "178650101U": {"tail": "N837PA", "block": "10"}, "178650101V": {"tail": "N843PA", "block": "10"}, "178650101W": {"tail": "N844PA", "block": "10"}, "178650101X": {"tail": "N856PA", "block": "10"}, "1786501020": {"tail": "N869PA", "block": "10"}, "1786501021": {"tail": "N880PA", "block": "20"}, "1786501022": {"tail": "N882PA", "block": "20"}, "1786501023": {"tail": "N885PA", "block": "20"}, "1786501024": {"tail": "N886PA", "block": "20"}, "1786501025": {"tail": "N889PA", "block": "20"}, "1786501026": {"tail": "N904PA", "block": "20"}, "1786501027": {"tail": "N913PA", "block": "20"}, "1786501028": {"tail": "N919PA", "block": "20"}, "1786501029": {"tail": "N932PA", "block": "20"}, "178650102A": {"tail": "N933PA", "block": "20"}, "178650102B": {"tail": "N934PA", "block": "20"}, "178650102C": {"tail": "N936PA", "block": "20"}, "178650102D": {"tail": "N937PA", "block": "20"}, "178650102E": {"tail": "N938PA", "block": "20"}, "178650102F": {"tail": "N940PA", "block": "20"}, "178650102G": {"tail": "N941PA", "block": "20"}, "178650102H": {"tail": "N942PA", "block": "20"}, "178650102J": {"tail": "N945PA", "block": "20"}, "178650102K": {"tail": "N946PA", "block": "20"}, "178650102L": {"tail": "N947PA", "block": "20"}, "178650102M": {"tail": "N949PA", "block": "20"}, "178650102N": {"tail": "N951PA", "block": "20"}, "178650102P": {"tail": "", "block": ""}, "178650102Q": {"tail": "N952PA", "block": "20"}, "178650102R": {"tail": "N954PA", "block": "20"}, "178650102S": {"tail": "N957PA", "block": "20"}, "178650102T": {"tail": "N958PA", "block": "20"}, "178650102U": {"tail": "N960PA", "block": "20"}, "178650102V": {"tail": "N962PA", "block": "20"}, "178650102W": {"tail": "N963PA", "block": "20"}, "178650102X": {"tail": "N964PA", "block": "20"}, "1786501030": {"tail": "N965PA", "block": "20"}, "1786501031": {"tail": "N966PA", "block": "20"}, "1786501032": {"tail": "N967PA", "block": "20"}, "1786501033": {"tail": "N968PA", "block": "20"}, "1786501034": {"tail": "N970PA", "block": "20"}, "1786501035": {"tail": "N975PA", "block": "20"}, "1786501036": {"tail": "N978PA", "block": "20"}, "1786501037": {"tail": "N979PA", "block": "20"}, "1786501038": {"tail": "N103PA", "block": "20"}, "1786501039": {"tail": "N108PA", "block": "20"}, "178650103A": {"tail": "N125PA", "block": "20"}, "178650103B": {"tail": "N126PA", "block": "20"}, "178650103C": {"tail": "N134PA", "block": "20"}, "178650103D": {"tail": "N142PA", "block": "20"}, "178650103E": {"tail": "N143PA", "block": "20"}, "178650103F": {"tail": "N239PA", "block": "20"}, "178650103G": {"tail": "N240PA", "block": "20"}, "178650103H": {"tail": "N243PA", "block": "20"}, "178650103J": {"tail": "N247PA", "block": "20"}, "178650103K": {"tail": "N249PA", "block": "20"}, "178650103L": {"tail": "N259PA", "block": "20"}, "178650103M": {"tail": "N263PA", "block": "20"}, "178650103N": {"tail": "N267PA", "block": "20"}, "178650103P": {"tail": "N139PA", "block": "20"}, "178650103Q": {"tail": "N304PA", "block": "20"}, "178650103R": {"tail": "N311PA", "block": "20"}, "178650103S": {"tail": "N319PA", "block": "20"}, "178650103T": {"tail": "N327PA", "block": "20"}, "178650103U": {"tail": "N328PA", "block": "20"}, "178650103V": {"tail": "N347PA", "block": "20"}, "178650103W": {"tail": "N351PA", "block": "20"}, "178650103X": {"tail": "N362PA", "block": "20"}, "1786501040": {"tail": "N363PA", "block": "20"}, "1786501041": {"tail": "N364PA", "block": "20"}, "1786501042": {"tail": "N381PA", "block": "20"}, "1786501043": {"tail": "N388PA", "block": "20"}, "1786501044": {"tail": "N390PA", "block": "20"}, "1786501045": {"tail": "N391PA", "block": "20"}, "1786501046": {"tail": "N392PA", "block": "20"}, "1786501047": {"tail": "N394PA", "block": "20"}, "1786501048": {"tail": "N395PA", "block": "20"}, "1786501049": {"tail": "N396PA", "block": "20"}, "178650104A": {"tail": "N408PA", "block": "20"}, "178650104B": {"tail": "N413PA", "block": "20"}, "178650104C": {"tail": "N420PA", "block": "20"}, "178650104D": {"tail": "N422PA", "block": "20"}, "178650104E": {"tail": "N439PA", "block": "20"}, "178650104F": {"tail": "N463PA", "block": "20"}, "178650104G": {"tail": "N466PA", "block": "20"}, "178650104H": {"tail": "N467PA", "block": "20"}, "178650104J": {"tail": "N469PA", "block": "20"}, "178650104K": {"tail": "N470PA", "block": "20"}, "178650104L": {"tail": "N475PA", "block": "30"}, "178650104M": {"tail": "N479PA", "block": "30"}, "178650104N": {"tail": "N480PA", "block": "30"}, "178650104P": {"tail": "N482PA", "block": "30"}, "178650104Q": {"tail": "N483PA", "block": "30"}, "178650104R": {"tail": "N488PA", "block": "20"}, "178650104S": {"tail": "N494PA", "block": "20"}, "178650104T": {"tail": "N496PA", "block": "20"}, "178650104U": {"tail": "N498PA", "block": "20"}, "178650104V": {"tail": "N516PA", "block": "20"}, "178650104W": {"tail": "N523PA", "block": "20"}, "178650104X": {"tail": "N526PA", "block": "20"}, "1786501050": {"tail": "N530PA", "block": "20"}, "1786501051": {"tail": "N538PA", "block": "20"}, "1786501052": {"tail": "N553PA", "block": "20"}, "1786501053": {"tail": "N563PA", "block": "20"}, "1786501054": {"tail": "N571PA", "block": "20"}, "1786501055": {"tail": "N572PA", "block": "20"}, "1786501056": {"tail": "N578PA", "block": "20"}, "1786501057": {"tail": "N580PA", "block": "20"}, "1786501058": {"tail": "N586PA", "block": "20"}, "1786501059": {"tail": "N588PA", "block": "20"}, "178650105A": {"tail": "N589PA", "block": "20"}, "178650105B": {"tail": "N590PA", "block": "20"}, "178650105C": {"tail": "", "block": "20"}, "178650105D": {"tail": "N592PA", "block": "20"}, "178650105E": {"tail": "N593PA", "block": "20"}, "178650105F": {"tail": "N596PA", "block": "20"}, "178650105G": {"tail": "N597PA", "block": "20"}, "178650105H": {"tail": "N617PA", "block": "20"}, "178650105J": {"tail": "N636PA", "block": "20"}, "178650105K": {"tail": "N647PA", "block": "20"}, "178650105L": {"tail": "N656PA", "block": "20"}, "178650105M": {"tail": "N692PA", "block": "20"}, "178650105N": {"tail": "N716PA", "block": "20"}, "178650105P": {"tail": "N721PA", "block": "20"}, "178650105Q": {"tail": "N723PA", "block": "20"}, "178650105R": {"tail": "N724PA", "block": "20"}, "178650105S": {"tail": "N733PA", "block": "20"}, "178650105T": {"tail": "N738PA", "block": "20"}, "178650105U": {"tail": "", "block": "20"}, "178650105V": {"tail": "N765PA", "block": "20"}, "178650105W": {"tail": "N766PA", "block": "20"}, "178650105X": {"tail": "N769PA", "block": "20"}, "1786501060": {"tail": "N774PA", "block": "20"}, "1786501061": {"tail": "N775PA", "block": "20"}, "1786501062": {"tail": "", "block": "20"}, "1786501063": {"tail": "", "block": "20"}, "1786501064": {"tail": "N983PA", "block": "20"}, "1786501065": {"tail": "N985PA", "block": "20"}, "1786501066": {"tail": "N986PA", "block": "20"}, "1786501067": {"tail": "N987PA", "block": "20"}, "1786501068": {"tail": "N992PA", "block": "20"}, "1786501069": {"tail": "N993PA", "block": "20"}, "178650106A": {"tail": "N994PA", "block": "20"}, "178650106B": {"tail": "N995PA", "block": "20"}, "178650106C": {"tail": "N217PA", "block": "20"}, "178650106D": {"tail": "N342PA", "block": "20"}, "178650106E": {"tail": "N361PA", "block": "20"}, "178650106F": {"tail": "N367PA", "block": "20"}, "178650106G": {"tail": "N382PA", "block": "20"}, "178650106H": {"tail": "N409PA", "block": "20"}, "178650106J": {"tail": "N489PA", "block": "20"}, "178650106K": {"tail": "N522PA", "block": "20"}, "178650106L": {"tail": "N535PA N955PA", "block": "20"}, "178650106M": {"tail": "N541PA N981PA", "block": "20"}, "178650106N": {"tail": "N542PA", "block": "20"}, "178650106P": {"tail": "N564PA", "block": "20"}, "178650106Q": {"tail": "N566PA", "block": "20"}, "178650106R": {"tail": "N642PA", "block": "20"}, "178650106S": {"tail": "N645PA", "block": "20"}, "178650106T": {"tail": "N664PA", "block": "20"}, "178650106U": {"tail": "N668PA", "block": "20"}, "178650106V": {"tail": "N669PA", "block": "20"}, "178650106W": {"tail": "N670PA", "block": "20"}, "178650106X": {"tail": "N697PA", "block": "20"}, "1786501070": {"tail": "N699PA", "block": "20"}, "1786501071": {"tail": "N703PA", "block": "20"}, "1786501072": {"tail": "N790PA", "block": "20"}, "1786501073": {"tail": "N867PA", "block": "20"}, "1786501074": {"tail": "N889PA", "block": "20"}, "1786501075": {"tail": "N922PA", "block": "20"}, "1786501076": {"tail": "N923PA", "block": "20"}, "1786501077": {"tail": "N740BG", "block": "20"}, "1786501078": {"tail": "N741BG", "block": "20"}, "1786501079": {"tail": "N139AM", "block": "20"}, "178650107A": {"tail": "N145AM", "block": "20"}, "178650107B": {"tail": "N165AM", "block": "20"}, "178650107C": {"tail": "N205AM", "block": "20"}, "178650107D": {"tail": "N268AM", "block": "20"}, "178650107E": {"tail": "N345AM", "block": "20"}, "178650107F": {"tail": "N361AM", "block": "20"}, "178650107G": {"tail": "N362AM", "block": "20"}, "178650107H": {"tail": "N363AM", "block": "20"}, "178650107J": {"tail": "N380AM", "block": "20"}, "178650107K": {"tail": "N387AM", "block": "20"}, "178650107L": {"tail": "N393AM", "block": "20"}, "178650107M": {"tail": "N401AM", "block": "20"}, "178650107N": {"tail": "N408AM", "block": "20"}, "178650107P": {"tail": "N423AM", "block": "20"}, "178650107Q": {"tail": "N443AM", "block": "20"}, "178650107R": {"tail": "N454AM", "block": "20"}, "178650107S": {"tail": "N460AM", "block": "20"}, "178650107T": {"tail": "N461AM", "block": "20"}, "178650107U": {"tail": "N462AM", "block": "20"}, "178650107V": {"tail": "N472AM", "block": "20"}, "178650107W": {"tail": "N476AM", "block": "20"}, "178650107X": {"tail": "N483AM", "block": "20"}, "1786501080": {"tail": "N484AM", "block": "20"}, "1786501081": {"tail": "N491AM", "block": "20"}, "1786501082": {"tail": "N493AM", "block": "20"}, "1786501083": {"tail": "N495AM", "block": "20"}, "1786501084": {"tail": "N496AM N571AM", "block": "20"}, "1786501085": {"tail": "N497AM", "block": "20"}, "1786501086": {"tail": "N498AM", "block": "20"}, "1786501087": {"tail": "N506AM", "block": "20"}, "1786501088": {"tail": "N526AM", "block": "20"}, "1786501089": {"tail": "N546AM", "block": "20"}, "178650108A": {"tail": "N548AM", "block": "20"}, "178650108B": {"tail": "N557AM", "block": "20"}, "178650108C": {"tail": "N558AM", "block": "20"}, "178650108D": {"tail": "N560AM", "block": "20"}, "178650108E": {"tail": "N564AM", "block": "20"}, "178650108F": {"tail": "N572AM", "block": "20"}, "178650108G": {"tail": "N577AM", "block": "20"}, "178650108H": {"tail": "N580AM", "block": "20"}, "178650108J": {"tail": "N581AM", "block": "20"}, "178650108K": {"tail": "N583AM", "block": "20"}, "178650108L": {"tail": "N590AM", "block": "20"}, "178650108M": {"tail": "N591AM", "block": "20"}, "178650108N": {"tail": "N592AM", "block": "20"}, "178650108P": {"tail": "N595AM", "block": "20"}, "178650108Q": {"tail": "N596AM", "block": "20"}, "178650108R": {"tail": "N597AM", "block": "20"}, "178650108S": {"tail": "N598AM", "block": "20"}, "178650108T": {"tail": "N603AM", "block": "20"}, "178650108U": {"tail": "N605AM", "block": "20"}, "178650108V": {"tail": "N609AM", "block": "20"}, "178650108W": {"tail": "N627AM", "block": "20"}, "178650108X": {"tail": "N636AM", "block": "20"}, "1786501090": {"tail": "N640AM", "block": "20"}, "1786501091": {"tail": "N641AM", "block": "20"}, "1786501092": {"tail": "N647AM", "block": "20"}, "1786501093": {"tail": "N648AM", "block": "20"}, "1786501094": {"tail": "N655AM", "block": "20"}, "1786501095": {"tail": "N661AM", "block": "20"}, "1786501096": {"tail": "N672AM", "block": "20"}, "1786501097": {"tail": "N679AM", "block": "20"}, "1786501098": {"tail": "N681AM", "block": "20"}, "1786501099": {"tail": "N689AM", "block": "20"}, "178650109A": {"tail": "N692AM", "block": "20"}, "178650109B": {"tail": "N698AM", "block": "20"}, "178650109C": {"tail": "N705AM", "block": "20"}, "178650109D": {"tail": "N707AM", "block": "20"}, "178650109E": {"tail": "N746AM", "block": "20"}, "178650109F": {"tail": "N763AM", "block": "20"}, "178650109G": {"tail": "N765AM", "block": "20"}, "178650109H": {"tail": "N774AM", "block": "20"}, "178650109J": {"tail": "N798AM", "block": "20"}, "178650109K": {"tail": "N801AM", "block": "20"}, "178650109L": {"tail": "N806AM", "block": "20"}, "178650109M": {"tail": "N814AM", "block": "20"}, "178650109N": {"tail": "N849AM", "block": "20"}, "178650109P": {"tail": "N851AM", "block": "20"}, "178650109Q": {"tail": "N861AM", "block": "20"}, "178650109R": {"tail": "N862AM", "block": "20"}, "178650109S": {"tail": "N863AM", "block": "20"}, "178650109T": {"tail": "N873AM", "block": "20"}, "178650109U": {"tail": "N876AM", "block": "20"}, "178650109V": {"tail": "N886AM", "block": "20"}, "178650109W": {"tail": "N890AM", "block": "20"}, "178650109X": {"tail": "N893AM", "block": "20"}, "1786501100": {"tail": "N897AM", "block": "20"}, "1786501101": {"tail": "N906AM", "block": "20"}, "1786501102": {"tail": "N913AM", "block": "20"}, "1786501103": {"tail": "N919AM", "block": "20"}, "1786501104": {"tail": "N934AM", "block": "20"}, "1786501105": {"tail": "N943AM", "block": "20"}, "1786501106": {"tail": "N949AM", "block": "20"}, "1786501107": {"tail": "N952AM", "block": "20"}, "1786501108": {"tail": "N954AM", "block": "20"}, "1786501109": {"tail": "N968AM", "block": "20"}, "178650111A": {"tail": "N702BG", "block": "20"}, "178650111B": {"tail": "N705BG", "block": "20"}, "178650111C": {"tail": "N706BG", "block": "20"}, "178650111D": {"tail": "N660BG", "block": "20"}, "178650111E": {"tail": "N662BG", "block": "20"}, "178650111F": {"tail": "N673BG", "block": "20"}, "178650111G": {"tail": "N670BG", "block": "20"}, "178650111H": {"tail": "N681BG", "block": "20"}, "178650111J": {"tail": "N682BG", "block": "20"}, "178650111K": {"tail": "N679BG", "block": "20"}, "178650111L": {"tail": "N693BG", "block": "20"}, "178650111M": {"tail": "N694BG", "block": "20"}, "178650111N": {"tail": "N687BG", "block": "20"}, "178650111P": {"tail": "N664BG", "block": "20"}, "178650111Q": {"tail": "N570BG", "block": "20"}, "178650111R": {"tail": "N633BG", "block": "20"}, "178650111S": {"tail": "N594BG", "block": "20"}, "178650111T": {"tail": "N574BG", "block": "20"}, "178650111U": {"tail": "N635BG", "block": "20"}, "178650111V": {"tail": "N583BG", "block": "20"}, "178650111W": {"tail": "N644BG", "block": "20"}, "178650111X": {"tail": "N612BG", "block": "20"}, "1786501111": {"tail": "N586BG", "block": "20"}, "1786501112": {"tail": "N646BG", "block": "20"}, "1786501113": {"tail": "N593BG", "block": "20"}, "1786501114": {"tail": "N573BG", "block": "20"}, "1786501115": {"tail": "N634BG", "block": "20"}, "1786501116": {"tail": "N596BG", "block": "20"}, "1786501117": {"tail": "N576BG", "block": "20"}, "1786501118": {"tail": "N611BG", "block": "20"}, "1786501119": {"tail": "N584BG", "block": "20"}, "178650112A": {"tail": "N645BG", "block": "20"}, "178650112B": {"tail": "N615BG", "block": "20"}, "178650112C": {"tail": "N587BG", "block": "20"}, "178650112D": {"tail": "N665BG", "block": "20"}, "178650112E": {"tail": "N655BG", "block": "20"}, "178650112F": {"tail": "N969BG", "block": "20"}, "178650112G": {"tail": "N685BG", "block": "20"}, "178650112H": {"tail": "N686BG", "block": "20"}, "178650112J": {"tail": "N674BG", "block": "20"}, "178650112K": {"tail": "N667BG", "block": "20"}, "178650112L": {"tail": "N657BG", "block": "20"}, "178650112M": {"tail": "N659BG", "block": "20"}, "178650112N": {"tail": "N683BG", "block": "20"}, "178650112P": {"tail": "B675BG", "block": "20"}, "178650112Q": {"tail": "N668BG", "block": "20"}, "178650112R": {"tail": "N669BG", "block": "20"}, "178650112S": {"tail": "N695BG", "block": "20"}, "178650112T": {"tail": "N684BG", "block": "20"}, "178650112U": {"tail": "N676BG", "block": "20"}, "178650112V": {"tail": "N677BG", "block": "20"}, "178650112W": {"tail": "N568BG", "block": "20"}, "178650112X": {"tail": "N630BG", "block": "20"}, "1786501120": {"tail": "N582BG", "block": "20"}, "1786501121": {"tail": "N643BG", "block": "20"}, "1786501122": {"tail": "N581BG", "block": "20"}, "1786501123": {"tail": "N641BG", "block": "20"}, "1786501124": {"tail": "N591BG", "block": "20"}, "1786501125": {"tail": "N653BG", "block": "20"}, "1786501126": {"tail": "N590BG", "block": "20"}, "1786501127": {"tail": "N652BG", "block": "20"}, "1786501128": {"tail": "N60BG", "block": "20"}, "1786501129": {"tail": "N607BG", "block": "20"}, "178650113A": {"tail": "N569BG", "block": "30"}, "178650113B": {"tail": "N631BG", "block": "30"}, "178650113C": {"tail": "N597BG", "block": "30"}, "178650113D": {"tail": "N578BG", "block": "30"}, "178650113E": {"tail": "N637BG", "block": "30"}, "178650113F": {"tail": "N589BG", "block": "30"}, "178650113G": {"tail": "N651BG", "block": "30"}, "178650113H": {"tail": "N621BG", "block": "30"}, "178650113J": {"tail": "N588BG", "block": "30"}, "178650113K": {"tail": "N648BG", "block": "30"}, "178650113L": {"tail": "N604BG", "block": "30"}, "178650113M": {"tail": "N636BG", "block": "30"}, "178650113N": {"tail": "N600BG", "block": "30"}, "178650113P": {"tail": "N567BG", "block": "30"}, "178650113Q": {"tail": "N629BG", "block": "30"}, "178650113R": {"tail": "N647BG", "block": "30"}, "178650113S": {"tail": "N627BG", "block": "30"}, "178650113T": {"tail": "N579BG", "block": "30"}, "178650113U": {"tail": "N638BG", "block": "30"}, "178650113V": {"tail": "N194BG", "block": "30"}, "178650113W": {"tail": "N274BG", "block": "30"}, "178650113X": {"tail": "N388BG", "block": "30"}, "1786501130": {"tail": "N467BG", "block": "20"}, "1786501131": {"tail": "N511BG", "block": "20"}, "1786501132": {"tail": "N233BG", "block": "20"}, "1786501133": {"tail": "N297BG", "block": "20"}, "1786501134": {"tail": "N402BG", "block": "20"}, "1786501135": {"tail": "N479BG", "block": "20"}, "1786501136": {"tail": "N532BG", "block": "20"}, "1786501137": {"tail": "N253BG", "block": "20"}, "1786501138": {"tail": "N316BG", "block": "20"}, "1786501139": {"tail": "N420BG", "block": "20"}, "178650114A": {"tail": "N487BG", "block": "20"}, "178650114B": {"tail": "N543BG", "block": "20"}, "178650114C": {"tail": "N264BG", "block": "20"}, "178650114D": {"tail": "N376BG", "block": "20"}, "178650114E": {"tail": "N437BG", "block": "20"}, "178650114F": {"tail": "N498BG", "block": "20"}, "178650114G": {"tail": "N558BG", "block": "20"}, "178650114H": {"tail": "N215BG", "block": "20"}, "178650114J": {"tail": "N288BG", "block": "20"}, "178650114K": {"tail": "N394BG", "block": "20"}, "178650114L": {"tail": "N483BG", "block": "20"}, "178650114M": {"tail": "N537BG", "block": "20"}, "178650114N": {"tail": "N246BG", "block": "20"}, "178650114P": {"tail": "N304BG", "block": "20"}, "178650114Q": {"tail": "N432BG", "block": "20"}, "178650114R": {"tail": "N493BG", "block": "20"}, "178650114S": {"tail": "N549BG", "block": "20"}, "178650114T": {"tail": "N257BG", "block": "20"}, "178650114U": {"tail": "N329BG", "block": "20"}, "178650114V": {"tail": "N463BG", "block": "20"}, "178650114W": {"tail": "N507BG", "block": "20"}, "178650114X": {"tail": "N565BG", "block": "20"}, "1786501140": {"tail": "N269BG", "block": "20"}, "1786501141": {"tail": "N382BG", "block": "20"}, "1786501142": {"tail": "N472BG", "block": "20"}, "1786501143": {"tail": "N523BG", "block": "20"}, "1786501144": {"tail": "N147BG", "block": "20"}, "1786501145": {"tail": "N271BG", "block": "20"}, "1786501146": {"tail": "N386BG", "block": "20"}, "1786501147": {"tail": "N465BG", "block": "20"}, "1786501148": {"tail": "N509BG", "block": "20"}, "1786501149": {"tail": "N229BG", "block": "20"}, "178650115A": {"tail": "N295BG", "block": "20"}, "178650115B": {"tail": "N398BG", "block": "20"}, "178650115C": {"tail": "N476BG", "block": "20"}, "178650115D": {"tail": "N528BG", "block": "20"}, "178650115E": {"tail": "N249BG", "block": "20"}, "178650115F": {"tail": "N308BG", "block": "20"}, "178650115G": {"tail": "N415BG", "block": "20"}, "178650115H": {"tail": "N485BG", "block": "20"}, "178650115J": {"tail": "N541BG", "block": "20"}, "178650115K": {"tail": "N261BG", "block": "20"}, "178650115L": {"tail": "N374BG", "block": "20"}, "178650115M": {"tail": "N435BG", "block": "20"}, "178650115N": {"tail": "N495BG", "block": "20"}, "178650115P": {"tail": "N554BG", "block": "20"}, "178650115Q": {"tail": "N129BG", "block": "20"}, "178650115R": {"tail": "N270BG", "block": "20"}, "178650115S": {"tail": "N384BG", "block": "20"}, "178650115T": {"tail": "N464BG", "block": "20"}, "178650115U": {"tail": "N508BG", "block": "20"}, "178650115V": {"tail": "N223BG", "block": "20"}, "178650115W": {"tail": "N289BG", "block": "20"}, "178650115X": {"tail": "N395BG", "block": "20"}, "1786501150": {"tail": "N475BG", "block": "20"}, "1786501151": {"tail": "N526BG", "block": "20"}, "1786501152": {"tail": "N248BG", "block": "20"}, "1786501153": {"tail": "N306BG", "block": "20"}, "1786501154": {"tail": "N413BG", "block": "20"}, "1786501155": {"tail": "N484BG", "block": "20"}, "1786501156": {"tail": "N538BG", "block": "20"}, "1786501157": {"tail": "N258BG", "block": "20"}, "1786501158": {"tail": "N330BG", "block": "20"}, "1786501159": {"tail": "N433BG", "block": "20"}, "178650116A": {"tail": "N494BG", "block": "20"}, "178650116B": {"tail": "N553BG", "block": "20"}, "178650116C": {"tail": "N196BG", "block": "20"}, "178650116D": {"tail": "N275BG", "block": "20"}, "178650116E": {"tail": "N389BG", "block": "20"}, "178650116F": {"tail": "N468BG", "block": "20"}, "178650116G": {"tail": "N513BG", "block": "20"}, "178650116H": {"tail": "N236BG", "block": "20"}, "178650116J": {"tail": "N298BG", "block": "20"}, "178650116K": {"tail": "N405BG", "block": "20"}, "178650116L": {"tail": "N480BG", "block": "20"}, "178650116M": {"tail": "N534BG", "block": "20"}, "178650116N": {"tail": "N254BG", "block": "20"}, "178650116P": {"tail": "N319BG", "block": "20"}, "178650116Q": {"tail": "N423BG", "block": "20"}, "178650116R": {"tail": "N488BG", "block": "20"}, "178650116S": {"tail": "N544BG", "block": "20"}, "178650116T": {"tail": "N265BG", "block": "20"}, "178650116U": {"tail": "N379BG", "block": "20"}, "178650116V": {"tail": "N438BG", "block": "20"}, "178650116W": {"tail": "N499BG", "block": "20"}, "178650116X": {"tail": "N561BG", "block": "20"}, "1786501160": {"tail": "N197BG", "block": "20"}, "1786501161": {"tail": "N278BG", "block": "20"}, "1786501162": {"tail": "N392BG", "block": "20"}, "1786501163": {"tail": "N469BG", "block": "20"}, "1786501164": {"tail": "N514BG", "block": "20"}, "1786501165": {"tail": "N243BG", "block": "20"}, "1786501166": {"tail": "N301BG", "block": "20"}, "1786501167": {"tail": "N406BG", "block": "20"}, "1786501168": {"tail": "N481BG", "block": "20"}, "1786501169": {"tail": "N535BG", "block": "20"}, "178650117A": {"tail": "N255BG", "block": "20"}, "178650117B": {"tail": "N326BG", "block": "20"}, "178650117C": {"tail": "N428BG", "block": "20"}, "178650117D": {"tail": "N489BG", "block": "20"}, "178650117E": {"tail": "N546BG", "block": "20"}, "178650117F": {"tail": "N266BG", "block": "20"}, "178650117G": {"tail": "N380BG", "block": "20"}, "178650117H": {"tail": "N440BG", "block": "20"}, "178650117J": {"tail": "N503BG", "block": "20"}, "178650117K": {"tail": "N563BG", "block": "20"}, "178650117L": {"tail": "N214BG", "block": "20"}, "178650117M": {"tail": "N302BG", "block": "20"}, "178650117N": {"tail": "N410BG", "block": "20"}, "178650117P": {"tail": "N482BG", "block": "20"}, "178650117Q": {"tail": "N536BG", "block": "20"}, "178650117R": {"tail": "N244BG", "block": "20"}, "178650117S": {"tail": "N328BG", "block": "20"}, "178650117T": {"tail": "N430BG", "block": "20"}, "178650117U": {"tail": "N490BG", "block": "20"}, "178650117V": {"tail": "N547BG", "block": "20"}, "178650117W": {"tail": "N256BG", "block": "20"}, "178650117X": {"tail": "N381BG", "block": "20"}, "1786501170": {"tail": "N452BG", "block": "20"}, "1786501171": {"tail": "N506BG", "block": "20"}, "1786501172": {"tail": "N564BG", "block": "20"}, "1786501173": {"tail": "N279BG", "block": "20"}, "1786501174": {"tail": "N393BG", "block": "20"}, "1786501175": {"tail": "N470BG", "block": "20"}, "1786501176": {"tail": "N517BG", "block": "20"}, "1786501177": {"tail": "N162BG", "block": "20"}, "1786501178": {"tail": "N273BG", "block": "20"}, "1786501179": {"tail": "N387BG", "block": "20"}, "178650118A": {"tail": "N466BG", "block": "20"}, "178650118B": {"tail": "N510BG", "block": "20"}, "178650118C": {"tail": "N230BG", "block": "20"}, "178650118D": {"tail": "N296BG", "block": "20"}, "178650118E": {"tail": "N401BG", "block": "20"}, "178650118F": {"tail": "N478BG", "block": "20"}, "178650118G": {"tail": "N531BG", "block": "20"}, "178650118H": {"tail": "N252BG", "block": "20"}, "178650118J": {"tail": "N313BG", "block": "20"}, "178650118K": {"tail": "N417BG", "block": "20"}, "178650118L": {"tail": "N486BG", "block": "20"}, "178650118M": {"tail": "N542BG", "block": "20"}, "178650118N": {"tail": "N263BG", "block": "20"}, "178650118P": {"tail": "N375BG", "block": "20"}, "178650118Q": {"tail": "N436BG", "block": "20"}, "178650118R": {"tail": "N496BG", "block": "20"}, "178650118S": {"tail": "N557BG", "block": "20"}, "178650118T": {"tail": "N742BG", "block": "20"}, "178650118U": {"tail": "N743BG", "block": "20"}, "178650118V": {"tail": "", "block": "20"}, "178650118W": {"tail": "", "block": "20"}, "178650118X": {"tail": "", "block": "20"}, "1786501180": {"tail": "", "block": "20"}, "1786501181": {"tail": "", "block": "20"}, "1786501182": {"tail": "", "block": "20"}, "1786501183": {"tail": "", "block": "20"}, "1786501184": {"tail": "", "block": "20"}, "1786501185": {"tail": "", "block": "20"}, "1786501186": {"tail": "", "block": "20"}, "1786501187": {"tail": "", "block": "20"}, "1786501188": {"tail": "", "block": "20"}, "1786501189": {"tail": "", "block": "20"}}

# Highlight colours
FILL_AMBER = PatternFill("solid", fgColor="FFD966")
FILL_RED   = PatternFill("solid", fgColor="FF7575")
FILL_GREEN = PatternFill("solid", fgColor="C6EFCE")
FONT_BOLD  = Font(bold=True)


# =============================================================================
# HELPERS
# =============================================================================

def _file_row(parent, label, row_num, on_browse):
    """A labelled file-picker row. Returns the label widget showing filename."""
    ctk.CTkLabel(
        parent, text=label,
        font=ctk.CTkFont(size=11),
        text_color=("gray40", "gray55"),
        width=160, anchor="w",
    ).grid(row=row_num, column=0, padx=(0, 10), pady=6, sticky="w")

    frame = ctk.CTkFrame(parent)
    frame.grid(row=row_num, column=1, pady=6, sticky="ew")
    frame.grid_columnconfigure(0, weight=1)

    lbl = ctk.CTkLabel(
        frame, text="No file selected",
        font=ctk.CTkFont(size=12),
        text_color=("gray40", "gray55"),
        anchor="w",
    )
    lbl.grid(row=0, column=0, padx=12, pady=8, sticky="ew")

    ctk.CTkButton(
        frame, text="Browse…", width=90,
        command=on_browse,
    ).grid(row=0, column=1, padx=(0, 8), pady=6)

    return lbl


def _section_label(parent, text):
    ctk.CTkLabel(
        parent, text=text,
        font=ctk.CTkFont(size=10, weight="bold"),
        text_color=("gray40", "gray50"),
        anchor="w",
    ).pack(fill="x", padx=0, pady=(14, 4))


def _status_lbl(parent):
    lbl = ctk.CTkLabel(parent, text="", font=ctk.CTkFont(size=11),
                       text_color=("gray40", "gray55"))
    lbl.pack(pady=(4, 0))
    return lbl


def set_status(lbl, msg, color=("gray40", "gray55")):
    lbl.configure(text=msg, text_color=color)


# =============================================================================
# COMPARISON LOGIC
# =============================================================================

def compare_44807(our_bom_path: str, website_path: str, out_path: str):
    """
    Reads:
      our_bom_path  — our generated _BOM.xlsx (reads the '44807' sheet)
      website_path  — website workbook with columns:
                      Part Type | Serial/Lot | Part Number | QTY | ... | Description

    Writes a comparison workbook with three sheets:
      1. Our 44807      — verbatim copy of our 44807 sheet
      2. Website 44807  — verbatim copy of the website sheet
      3. Result         — per-PN comparison with colour-coded status
    """

    # ── Load our 44807 ────────────────────────────────────────────────────────
    our_wb = load_workbook(our_bom_path, data_only=True)
    if "44807" not in our_wb.sheetnames:
        raise ValueError("'44807' sheet not found in the BOM workbook.")
    our_ws = our_wb["44807"]

    our_rows    = list(our_ws.values)
    our_headers = [str(c or "").strip() for c in our_rows[0]]
    our_data    = [dict(zip(our_headers, row)) for row in our_rows[1:]]

    def oc(row, col):
        return str(row.get(col) or "").strip()

    # ── Load website workbook ─────────────────────────────────────────────────
    # Format: Part Type | Serial/Lot | Part Number | QTY | (blank) | Description | ...
    web_wb = load_workbook(website_path, data_only=True)
    web_ws = web_wb.active
    web_rows = list(web_ws.iter_rows(values_only=True))

    # Find header row (contains "Part Number")
    header_idx = 0
    for i, row in enumerate(web_rows):
        if any("Part Number" in str(c or "") for c in row):
            header_idx = i
            break
    web_headers = [str(c or "").strip() for c in web_rows[header_idx]]

    # Locate key columns by header name
    def col_idx(name):
        for i, h in enumerate(web_headers):
            if name.lower() in h.lower():
                return i
        return None

    ci_pn    = col_idx("Part Number")
    ci_sl    = col_idx("Serial/Lot")
    ci_qty   = col_idx("QTY")
    ci_desc  = col_idx("Description")
    ci_type  = col_idx("Part Type")

    if ci_pn is None:
        raise ValueError("Could not find 'Part Number' column in website workbook.")

    # Build lookup: pn (upper) → {ids: set, qty: int}
    from collections import defaultdict
    web_pn_ids = defaultdict(set)
    web_pn_qty = defaultdict(int)

    for row in web_rows[header_idx + 1:]:
        pn = str(row[ci_pn] or "").strip() if ci_pn is not None else ""
        sl = str(row[ci_sl] or "").strip() if ci_sl is not None else ""
        try:
            qty = int(row[ci_qty]) if ci_qty is not None and row[ci_qty] else 0
        except (ValueError, TypeError):
            qty = 0
        if pn:
            if sl:
                web_pn_ids[pn.upper()].add(sl)
            web_pn_qty[pn.upper()] += qty

    # ── Build result rows ─────────────────────────────────────────────────────
    RES_HEADERS = [
        "Primary PN", "Matched PN", "Description",
        "SN/LOT", "QTY Required", "QTY Actual (Ours)", "QTY on Website",
        "Our IDs", "Our Status",
        "On Website", "IDs Missing from Website", "Result",
    ]

    result_rows = []

    for r in our_data:
        primary    = oc(r, "Primary PN")
        matched    = oc(r, "Matched PN")
        desc       = oc(r, "Description")
        snlot      = oc(r, "SN/LOT")
        qty_req    = oc(r, "QTY Required")
        qty_act    = oc(r, "QTY Actual")
        our_ids    = oc(r, "IDs")
        our_status = oc(r, "Status")

        # Our IDs — strip lot suffixes like "(x3)"
        our_id_clean = {
            re.sub(r"\s*\(x\d+\)$", "", x.strip())
            for x in our_ids.splitlines() if x.strip()
        }

        # Try matched PN first, fall back to primary
        lookup_pn   = matched if matched else primary
        lookup_key  = lookup_pn.upper() if lookup_pn else ""
        on_website  = lookup_key in web_pn_ids
        web_ids     = web_pn_ids.get(lookup_key, set())
        web_qty     = web_pn_qty.get(lookup_key, 0)

        # IDs we have that aren't on the website yet
        missing_ids = our_id_clean - web_ids if our_id_clean else set()

        if not lookup_pn or our_status == "NOT FOUND":
            result = "NOT IN BOM"
        elif not on_website:
            result = "MISSING FROM WEBSITE"
        elif missing_ids:
            result = "PARTIAL — IDs MISSING"
        else:
            result = "OK"

        result_rows.append([
            primary, matched, desc, snlot,
            qty_req, qty_act, web_qty if on_website else "",
            our_ids, our_status,
            "YES" if on_website else "NO",
            "\n".join(sorted(missing_ids)),
            result,
        ])

    # ── Write output workbook ─────────────────────────────────────────────────
    wb = Workbook()

    # Sheet 1 — Our 44807
    ws1 = wb.active
    ws1.title = "Our 44807"
    for row in our_ws.iter_rows(values_only=True):
        ws1.append([v if v is not None else "" for v in row])

    # Sheet 2 — Website 44807 (verbatim)
    ws2 = wb.create_sheet("Website 44807")
    for row in web_ws.iter_rows(values_only=True):
        ws2.append([v if v is not None else "" for v in row])

    # Sheet 3 — Result
    ws3 = wb.create_sheet("Result")
    ws3.append(RES_HEADERS)

    for i, row_data in enumerate(result_rows, start=2):
        result_val = row_data[-1]
        ws3.append(row_data)
        if result_val == "OK":
            fill = FILL_GREEN
        elif result_val == "PARTIAL — IDs MISSING":
            fill = FILL_AMBER
        elif result_val in ("MISSING FROM WEBSITE", "NOT IN BOM"):
            fill = FILL_RED
        else:
            fill = None
        if fill:
            for col in range(1, len(RES_HEADERS) + 1):
                ws3.cell(row=i, column=col).fill = fill
                ws3.cell(row=i, column=col).font = FONT_BOLD

    wb.save(out_path)


# =============================================================================
# TAB 1 — REGISTRATION GENERATOR
# =============================================================================

class Tab1(ctk.CTkFrame):
    def __init__(self, parent):
        super().__init__(parent, fg_color="transparent")
        self.grid_columnconfigure(0, weight=1)
        self._export_path = None
        self._build()

    def _build(self):
        # Step 1
        _section_label(self, "STEP 1  —  Upload Export File")

        upload_frame = ctk.CTkFrame(self)
        upload_frame.pack(fill="x", pady=(0, 6))
        upload_frame.grid_columnconfigure(0, weight=1)

        self._file_label = ctk.CTkLabel(
            upload_frame, text="No file selected",
            font=ctk.CTkFont(size=12),
            text_color=("gray40", "gray55"), anchor="w",
        )
        self._file_label.grid(row=0, column=0, padx=16, pady=12, sticky="w")

        ctk.CTkButton(
            upload_frame, text="Browse…", width=100,
            command=self._browse,
        ).grid(row=0, column=1, padx=(0, 12), pady=8)

        self._btn_process = ctk.CTkButton(
            self, text="Process Export →",
            font=ctk.CTkFont(size=13, weight="bold"),
            height=42, state="disabled",
            command=self._process,
        )
        self._btn_process.pack(fill="x", pady=(0, 4))

        self._status1 = _status_lbl(self)

        # Step 2
        _section_label(self, "STEP 2  —  Review & Configure")

        info_frame = ctk.CTkFrame(self)
        info_frame.pack(fill="x", pady=(0, 6))
        info_frame.grid_columnconfigure((0, 1), weight=1)
        self._disp_serial = self._info_box(info_frame, "Parent Serial #", "—", 0)
        self._disp_tail   = self._info_box(info_frame, "Tail # (from PDF)", "—", 1)

        self._not_found_label = ctk.CTkLabel(
            self,
            text="⚠  Serial not found in PDF — enter tail number manually.",
            font=ctk.CTkFont(size=11),
            text_color=("#c97a00", "#f0a44a"),
        )

        fields = ctk.CTkFrame(self, fg_color="transparent")
        fields.pack(fill="x", pady=(0, 6))
        fields.grid_columnconfigure(1, weight=1)

        self._in_serial = self._field(fields, "Serial Number", 0)
        self._in_tail   = self._field(fields, "Tail Number",   1)

        ctk.CTkLabel(
            fields, text="Block",
            font=ctk.CTkFont(size=11),
            text_color=("gray40", "gray55"),
            width=130, anchor="w",
        ).grid(row=2, column=0, padx=(0, 12), pady=6, sticky="w")

        self._in_block = ctk.CTkComboBox(fields, values=["20", "30"], state="readonly", width=300)
        self._in_block.set("")
        self._in_block.grid(row=2, column=1, pady=6, sticky="ew")

        # Step 3
        _section_label(self, "STEP 3  —  Generate")

        self._btn_generate = ctk.CTkButton(
            self,
            text="↓   Generate",
            font=ctk.CTkFont(size=13, weight="bold"),
            height=46,
            fg_color=("#1a7acc", "#1a7acc"),
            hover_color=("#1560a8", "#1560a8"),
            command=self._generate,
        )
        self._btn_generate.pack(fill="x", pady=(0, 4))

        self._status_gen = _status_lbl(self)

        ctk.CTkButton(
            self, text="← Start Over",
            fg_color="transparent", border_width=1,
            text_color=("gray40", "gray60"),
            hover_color=("gray85", "gray25"),
            height=34,
            command=self._reset,
        ).pack(anchor="w", pady=(8, 0))

    def _info_box(self, parent, label, value, col):
        frame = ctk.CTkFrame(parent)
        frame.grid(row=0, column=col, padx=12, pady=12, sticky="ew")
        ctk.CTkLabel(frame, text=label, font=ctk.CTkFont(size=10),
                     text_color=("gray40", "gray55")).pack(anchor="w", padx=12, pady=(10,2))
        lbl = ctk.CTkLabel(frame, text=value, font=ctk.CTkFont(size=14, weight="bold"))
        lbl.pack(anchor="w", padx=12, pady=(0,10))
        return lbl

    def _field(self, parent, label, row):
        ctk.CTkLabel(parent, text=label, font=ctk.CTkFont(size=11),
                     text_color=("gray40", "gray55"), width=130, anchor="w",
                     ).grid(row=row, column=0, padx=(0,12), pady=6, sticky="w")
        e = ctk.CTkEntry(parent, width=300, font=ctk.CTkFont(size=13))
        e.grid(row=row, column=1, pady=6, sticky="ew")
        return e

    def _browse(self):
        path = filedialog.askopenfilename(
            title="Select export file",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")],
        )
        if path:
            self._export_path = path
            self._file_label.configure(text=os.path.basename(path),
                                       text_color=("gray20","gray90"))
            self._btn_process.configure(state="normal")
            set_status(self._status1, "")

    def _process(self):
        self._btn_process.configure(state="disabled", text="Processing…")
        set_status(self._status1, "Reading export file…")
        threading.Thread(target=self._process_worker, daemon=True).start()

    def _process_worker(self):
        try:
            df = pd.read_excel(self._export_path, dtype=str).fillna("")
            df["ProductNo"] = df["ProductNo"].str.strip()
            anchor = df[df["ProductNo"] == ANCHOR_PN]
            if anchor.empty:
                raise ValueError(f"Anchor PN \'{ANCHOR_PN}\' not found")
            parents = anchor["Parent Serial #"].str.strip().replace("", pd.NA).dropna()
            if parents.empty:
                raise ValueError("No Parent Serial # found")
            serial = parents.value_counts().idxmax()
            info   = SERIAL_MAP.get(serial, {})
            tail   = info.get("tail", "")
            block  = info.get("block", "")
            found  = bool(tail)
            self.after(0, lambda: self._populate(serial, tail, block, found))
        except Exception as e:
            self.after(0, lambda: (
                set_status(self._status1, f"Error: {e}", ("#c0392b","#e74c3c")),
                self._btn_process.configure(state="normal", text="Process Export →"),
            ))

    def _populate(self, serial, tail, block, found):
        self._disp_serial.configure(text=serial)
        self._disp_tail.configure(
            text=tail if tail else "Not found in PDF",
            text_color=("gray40","gray55") if not tail else ("gray10","gray90"),
        )
        self._in_serial.delete(0,"end"); self._in_serial.insert(0, serial)
        self._in_tail.delete(0,"end");   self._in_tail.insert(0, tail)
        if block in ("20","30"): self._in_block.set(block)
        if not found:
            self._not_found_label.pack(anchor="w", pady=(0,4))
        else:
            self._not_found_label.pack_forget()
        set_status(self._status1,
            f"✓  {serial}" + (f"  ·  {tail}" if tail else "  ·  tail not in PDF"),
            ("#1a8a4a","#4af0a4"))
        self._btn_process.configure(state="normal", text="Process Export →")

    def _generate(self):
        serial = self._in_serial.get().strip()
        tail   = self._in_tail.get().strip()
        block  = self._in_block.get().strip()
        if not serial:
            messagebox.showerror("Missing field", "Serial Number is required."); return
        if not block:
            messagebox.showerror("Missing field", "Please select a Block."); return

        out_dir = filedialog.askdirectory(title="Select output folder")
        if not out_dir: return

        self._btn_generate.configure(state="disabled", text="Generating…")
        set_status(self._status_gen, "Running BOM processor…")
        threading.Thread(
            target=self._generate_worker,
            args=(serial, tail, block, out_dir),
            daemon=True,
        ).start()

    def _generate_worker(self, serial, tail, block, out_dir):
        try:
            import sys, importlib
            app_dir = os.path.dirname(os.path.abspath(__file__))
            if app_dir not in sys.path:
                sys.path.insert(0, app_dir)

            pdf_parser = importlib.import_module("pdf_parser")
            matcher_m  = importlib.import_module("matcher")
            writers_m  = importlib.import_module("writers")
            config_m   = importlib.import_module("config")
            utils_m    = importlib.import_module("utils")

            self.after(0, lambda: set_status(self._status_gen, "Parsing PDF template…"))
            parts, _ = pdf_parser.read_pdf(config_m.TEMPLATE_PATH)

            self.after(0, lambda: set_status(self._status_gen, "Loading export…"))
            df        = matcher_m.load_export(config_m.EXPORT_PATH)
            installed = {pn: grp for pn, grp in df.groupby("ProductNo")}

            self.after(0, lambda: set_status(self._status_gen, "Matching parts…"))
            sn_pool = {}
            results = []

            for part in parts:
                qty_req = part["qty"]
                if qty_req == 0: continue
                matched = matcher_m.find_match(part, installed)

                if matched:
                    rows     = installed[matched]
                    exp_desc = matcher_m.get_best_desc(rows)
                    snlot    = matcher_m.detect_snlot(rows)
                    if snlot == "SN":
                        ids     = matcher_m.allocate_sns(rows, qty_req, sn_pool, matched)
                        qty_act = len(ids)
                    else:
                        ids     = matcher_m.allocate_lots(rows)
                        qty_act = sum(
                            utils_m.to_int(m.group(1))
                            for s in ids if (m := re.search(r"\(x(\d+)\)", s))
                        )
                    status = "SATISFIED" if qty_act >= qty_req else "NOT SATISFIED"
                else:
                    old_pn = matcher_m.find_old_rev(part["preferred"], installed)
                    if old_pn:
                        rows     = installed[old_pn]
                        exp_desc = matcher_m.get_best_desc(rows)
                        snlot    = matcher_m.detect_snlot(rows)
                        ids      = (matcher_m.allocate_sns(rows, qty_req, sn_pool, old_pn)
                                    if snlot == "SN" else matcher_m.allocate_lots(rows))
                        qty_act  = len(ids)
                        matched  = old_pn
                        status   = "OLD REV"
                    else:
                        ids, qty_act, exp_desc, snlot = [], 0, "", "SN"
                        status = "NOT FOUND"

                results.append({
                    "pn": part["preferred"], "matched_pn": matched,
                    "desc": exp_desc or part["desc"], "snlot": snlot,
                    "qty_req": qty_req, "qty_act": qty_act,
                    "ids_text": "\n".join(ids), "status": status,
                })

            results.sort(key=lambda r: 0 if r["snlot"] == "SN" else 1)

            self.after(0, lambda: set_status(self._status_gen, "Writing files…"))
            base = os.path.join(out_dir, serial)
            writers_m.write_bom(f"{base}_BOM.xlsx", df, results, serial)
            writers_m.write_part_list(f"{base}_part_list.xlsx", results, serial)

            reg_path = f"{base}_registration.csv"
            row = list(CSV_DEFAULTS)
            def sc(name, val):
                if name in CSV_HEADERS: row[CSV_HEADERS.index(name)] = val
            sc("Serial Number", serial); sc("Tail Number", tail); sc("Block Number", block)
            with open(reg_path, "w", newline="") as f:
                import csv as _csv
                w = _csv.writer(f); w.writerow(CSV_HEADERS); w.writerow(row)

            sat = sum(1 for r in results if r["status"]=="SATISFIED")
            nf  = sum(1 for r in results if r["status"]=="NOT FOUND")
            olr = sum(1 for r in results if r["status"]=="OLD REV")
            msg = f"✓  3 files saved  ·  {sat} satisfied  ·  {nf} not found  ·  {olr} old rev"
            self.after(0, lambda: set_status(self._status_gen, msg, ("#1a8a4a","#4af0a4")))
            self.after(0, lambda: self._btn_generate.configure(state="normal", text="↓   Generate"))

        except Exception as e:
            err = str(e)
            self.after(0, lambda: set_status(self._status_gen, f"Error: {err}", ("#c0392b","#e74c3c")))
            self.after(0, lambda: self._btn_generate.configure(state="normal", text="↓   Generate"))

    def _reset(self):
        self._export_path = None
        self._file_label.configure(text="No file selected", text_color=("gray40","gray55"))
        self._btn_process.configure(state="disabled")
        set_status(self._status1, ""); set_status(self._status_gen, "")
        self._disp_serial.configure(text="—")
        self._disp_tail.configure(text="—", text_color=("gray10","gray90"))
        self._in_serial.delete(0,"end"); self._in_tail.delete(0,"end"); self._in_block.set("")
        self._not_found_label.pack_forget()


# =============================================================================
# TAB 2 — 44807 VS 44807
# =============================================================================

class Tab2(ctk.CTkFrame):
    def __init__(self, parent):
        super().__init__(parent, fg_color="transparent")
        self.grid_columnconfigure(0, weight=1)
        self._bom_path     = None
        self._website_path = None
        self._build()

    def _build(self):
        _section_label(self, "UPLOAD FILES")

        files_frame = ctk.CTkFrame(self, fg_color="transparent")
        files_frame.pack(fill="x", pady=(0, 6))
        files_frame.grid_columnconfigure(1, weight=1)

        self._lbl_bom = _file_row(
            files_frame, "Our BOM Workbook", 0,
            lambda: self._browse("bom"),
        )
        self._lbl_web = _file_row(
            files_frame, "Website Export", 1,
            lambda: self._browse("website"),
        )

        _section_label(self, "COMPARISON")

        ctk.CTkLabel(
            self,
            text=(
                "Compares every PN in our 44807 sheet against what the website has.\n"
                "Result sheet is colour coded:  🟢 OK  ·  🟡 IDs missing  ·  🔴 PN missing"
            ),
            font=ctk.CTkFont(size=11),
            text_color=("gray40","gray55"),
            justify="left",
            anchor="w",
        ).pack(fill="x", pady=(0, 10))

        self._btn_compare = ctk.CTkButton(
            self,
            text="↓   Compare & Generate",
            font=ctk.CTkFont(size=13, weight="bold"),
            height=46,
            fg_color=("#1a7acc","#1a7acc"),
            hover_color=("#1560a8","#1560a8"),
            command=self._compare,
        )
        self._btn_compare.pack(fill="x", pady=(0, 4))

        self._status = _status_lbl(self)

        ctk.CTkButton(
            self, text="← Clear",
            fg_color="transparent", border_width=1,
            text_color=("gray40","gray60"),
            hover_color=("gray85","gray25"),
            height=34,
            command=self._reset,
        ).pack(anchor="w", pady=(8, 0))

    def _browse(self, which):
        path = filedialog.askopenfilename(
            title="Select file",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")],
        )
        if not path: return
        name = os.path.basename(path)
        if which == "bom":
            self._bom_path = path
            self._lbl_bom.configure(text=name, text_color=("gray20","gray90"))
        else:
            self._website_path = path
            self._lbl_web.configure(text=name, text_color=("gray20","gray90"))
        set_status(self._status, "")

    def _compare(self):
        if not self._bom_path:
            messagebox.showerror("Missing file", "Please select our BOM workbook."); return
        if not self._website_path:
            messagebox.showerror("Missing file", "Please select the website export."); return

        out_path = filedialog.asksaveasfilename(
            title="Save comparison workbook",
            initialfile="44807_comparison.xlsx",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if not out_path: return

        self._btn_compare.configure(state="disabled", text="Comparing…")
        set_status(self._status, "Running comparison…")
        threading.Thread(
            target=self._compare_worker,
            args=(self._bom_path, self._website_path, out_path),
            daemon=True,
        ).start()

    def _compare_worker(self, bom_path, website_path, out_path):
        try:
            compare_44807(bom_path, website_path, out_path)
            msg = f"✓  Saved: {os.path.basename(out_path)}"
            self.after(0, lambda: set_status(self._status, msg, ("#1a8a4a","#4af0a4")))
        except Exception as e:
            err = str(e)
            self.after(0, lambda: set_status(self._status, f"Error: {err}", ("#c0392b","#e74c3c")))
        finally:
            self.after(0, lambda: self._btn_compare.configure(
                state="normal", text="↓   Compare & Generate"))

    def _reset(self):
        self._bom_path = self._website_path = None
        self._lbl_bom.configure(text="No file selected", text_color=("gray40","gray55"))
        self._lbl_web.configure(text="No file selected", text_color=("gray40","gray55"))
        set_status(self._status, "")


# =============================================================================
# MAIN APP WINDOW
# =============================================================================

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("MK30 Tools")
        self.geometry("660x780")
        self.resizable(False, False)
        self.grid_columnconfigure(0, weight=1)

        # Header
        hdr = ctk.CTkFrame(self, fg_color="transparent")
        hdr.grid(row=0, column=0, padx=30, pady=(28,8), sticky="ew")
        ctk.CTkLabel(hdr, text="Amazon Prime Air  ·  MK30",
                     font=ctk.CTkFont(size=11), text_color=("gray50","gray50")).pack(anchor="w")
        ctk.CTkLabel(hdr, text="MK30 Tools",
                     font=ctk.CTkFont(size=26, weight="bold")).pack(anchor="w", pady=(2,0))

        # Tab view
        tabs = ctk.CTkTabview(self, height=650)
        tabs.grid(row=1, column=0, padx=20, pady=(0,20), sticky="nsew")
        tabs.add("Registration Generator")
        tabs.add("44807 vs 44807")

        # Scrollable inner frames so content doesn't get clipped
        scroll1 = ctk.CTkScrollableFrame(tabs.tab("Registration Generator"),
                                         fg_color="transparent")
        scroll1.pack(fill="both", expand=True, padx=10, pady=6)
        scroll1.grid_columnconfigure(0, weight=1)
        Tab1(scroll1).pack(fill="both", expand=True)

        scroll2 = ctk.CTkScrollableFrame(tabs.tab("44807 vs 44807"),
                                         fg_color="transparent")
        scroll2.pack(fill="both", expand=True, padx=10, pady=6)
        scroll2.grid_columnconfigure(0, weight=1)
        Tab2(scroll2).pack(fill="both", expand=True)


if __name__ == "__main__":
    app = App()
    app.mainloop()