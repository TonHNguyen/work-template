"""
writers.py
----------
Writes the three output Excel files.

  {parent}_BOM.xlsx          — 4 sheets: full export, 44807 summary,
                               vehicle intake, missing components
  {parent}_part_list.xlsx    — vehicle intake only
  {parent}_registration.xlsx — registration placeholder
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill
import pandas as pd

from config import FILL_AMBER, FILL_RED, FONT_BOLD
from utils  import strip_lot_suffix


def _highlight_row(ws, row: int, fill: PatternFill, ncols: int):
    for col in range(1, ncols + 1):
        ws.cell(row=row, column=col).fill = fill
        ws.cell(row=row, column=col).font = FONT_BOLD


def _write_vehicle_intake(ws, results: list[dict], parent_serial: str):
    """Shared helper — writes the Vehicle intake rows into any worksheet."""
    ws.append(["Parent Serial #", "Serial/Lot"])
    for r in results:
        if r["status"] in ("SATISFIED", "NOT SATISFIED"):
            for line in r["ids_text"].splitlines():
                clean = strip_lot_suffix(line)
                if clean:
                    ws.append([parent_serial, clean])


def write_bom(path: str, export_df: pd.DataFrame,
              results: list[dict], parent_serial: str):
    wb = Workbook()

    # ── Sheet 1: Full export ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Combined as built"
    ws.append(list(export_df.columns))
    for row in export_df.itertuples(index=False):
        ws.append(list(row))

    # ── Sheet 2: 44807 summary ────────────────────────────────────────────────
    ws = wb.create_sheet("44807")
    ws.append(["Primary PN", "Matched PN", "Description",
               "SN/LOT", "QTY Required", "QTY Actual", "IDs", "Status"])
    for r in results:
        ws.append([r["pn"], r["matched_pn"], r["desc"],
                   r["snlot"], r["qty_req"], r["qty_act"],
                   r["ids_text"], r["status"]])

    # ── Sheet 3: Vehicle intake ───────────────────────────────────────────────
    ws = wb.create_sheet("Vehicle intake")
    _write_vehicle_intake(ws, results, parent_serial)

    # ── Sheet 4: Missing / problem parts ─────────────────────────────────────
    ws = wb.create_sheet("Missing Components")
    headers = ["Parent Serial #", "Required PN", "Found PN (Old Rev)",
               "Description", "SN/LOT", "QTY Required", "QTY Actual",
               "Missing Qty", "Status"]
    ws.append(headers)

    row_num = 2
    for r in results:
        if r["status"] == "SATISFIED":
            continue
        missing = max(0, r["qty_req"] - r["qty_act"])
        found   = r["matched_pn"] if r["status"] == "OLD REV" else ""
        ws.append([parent_serial, r["pn"], found, r["desc"],
                   r["snlot"], r["qty_req"], r["qty_act"], missing, r["status"]])
        fill = FILL_AMBER if r["status"] == "OLD REV" else FILL_RED
        _highlight_row(ws, row_num, fill, len(headers))
        row_num += 1

    wb.save(path)
    print("Saved:", path)


def write_part_list(path: str, results: list[dict], parent_serial: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Vehicle intake"
    _write_vehicle_intake(ws, results, parent_serial)
    wb.save(path)
    print("Saved:", path)


def write_registration(path: str, parent_serial: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Registration"
    ws.append(["Field", "Value"])
    ws.append(["Parent Serial #", parent_serial])
    ws.append(["Tail Number", ""])
    wb.save(path)
    print("Saved:", path)
