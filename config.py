"""
config.py
---------
All user-configurable settings in one place.
Change paths, anchor PN, rev-sensitive parts, or PN overrides here — nowhere else.
"""

from openpyxl.styles import PatternFill, Font

# ── File paths ────────────────────────────────────────────────────────────────
TEMPLATE_PATH = "templates/DOC-012743.pdf"
EXPORT_PATH   = "data/export_all_c80eb8e5ef3040379d61be5c46e8cd83.xlsx"
OUTPUT_DIR    = "outputs"

# ── Anchor PN ─────────────────────────────────────────────────────────────────
# Used to identify the drone's parent serial number from the export.
ANCHOR_PN = "LBL-F5-01"

# ── Rev-sensitive parts ───────────────────────────────────────────────────────
# Only these base PNs require an exact revision match.
# All other parts will use the highest available revision in the export.
# Add more base PNs here as needed e.g. {"147712", "150393"}
REV_SENSITIVE = {"147712"}

# ── Manual PN overrides ───────────────────────────────────────────────────────
# Parts that are NOT in the PDF yet but are valid replacements in the export.
# Format:  "pdf_pn" -> ["replacement_pn_1", "replacement_pn_2", ...]
#
# The matcher tries the PDF PN first, then falls through to these in order.
# To remove an override once the PDF is updated, just delete that line.
#
PN_OVERRIDES = {
    "159604-01": ["160275-02"],   # PROPELLER ASSY, T-MOTOR, 4-PLY, CENTER, CCW
    "159603-01": ["160274-02"],   # PROPELLER ASSY, T-MOTOR, 4-PLY, CENTER, CW
    "159602-01": ["160273-02"],   # PROPELLER ASSY, T-MOTOR, 4-PLY, OFF-CENTER, CCW
    "159601-01": ["160272-02"],   # PROPELLER ASSY, T-MOTOR, 4-PLY, OFF-CENTER, CW
}

# ── Cell highlight colours (Missing Components sheet) ─────────────────────────
FILL_AMBER = PatternFill("solid", fgColor="FFD966")   # OLD REV
FILL_RED   = PatternFill("solid", fgColor="FF7575")   # NOT FOUND
FONT_BOLD  = Font(bold=True)
