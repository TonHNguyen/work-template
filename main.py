from openpyxl import load_workbook, Workbook
import pandas as pd
import os
import re

TEMPLATE_PATH = "templates/template.xlsx"
EXPORT_PATH = "data/export_all_65c4618b186b4d068cc944cf7f28a71a.xlsx"
ALIASES_PATH = "data/aliases.xlsx"


def to_int_or_zero(v):
    if v is None:
        return 0
    if isinstance(v, str):
        s = v.strip()
        if s == "" or s.startswith("#"):
            return 0
        try:
            return int(float(s))
        except ValueError:
            return 0
    try:
        return int(v)
    except Exception:
        return 0


def to_float_or_zero(v):
    try:
        s = str(v).strip()
        if s == "" or s.startswith("#"):
            return 0.0
        return float(s)
    except Exception:
        return 0.0


def strip_lot_qty_suffix(s: str) -> str:
    # "BL01000001X5 (x4)" -> "BL01000001X5"
    return re.sub(r"\s*\(x\d+\)\s*$", "", str(s).strip())


def read_requirements():
    """
    Template tab 44807:
      A: Part number
      B: Description (template; used for non-missing fallback)
      C: Qty required (data_only=True uses cached values)
      E: SN/LOT
    Returns: list[(primary_pn, template_desc, qty_required, snlot)]
    """
    wb = load_workbook(TEMPLATE_PATH, keep_vba=True, data_only=True)
    ws = wb["44807"]

    requirements = []
    row = 2
    while True:
        pn = ws[f"A{row}"].value
        if pn is None or str(pn).strip() == "":
            break

        primary_pn = str(pn).strip()

        template_desc = ws[f"B{row}"].value
        template_desc = str(template_desc).strip() if template_desc else ""

        qty_required = to_int_or_zero(ws[f"C{row}"].value)

        sn_lot = ws[f"E{row}"].value
        sn_lot = str(sn_lot).strip().upper() if sn_lot else ""
        sn_lot = sn_lot if sn_lot in ("SN", "LOT") else "SN"

        requirements.append((primary_pn, template_desc, qty_required, sn_lot))
        row += 1

    return requirements


def load_export():
    return pd.read_excel(EXPORT_PATH, dtype=str).fillna("")


def load_aliases():
    """
    Aliases workbook:
      First sheet only
      Column A: PrimaryPartNo
      Columns B..N: alternates
    """
    df = pd.read_excel(ALIASES_PATH, sheet_name=0, dtype=str).fillna("")
    aliases = {}

    for _, row in df.iterrows():
        primary = str(row.iloc[0]).strip()
        if not primary:
            continue

        allowed = [primary]
        for v in row.iloc[1:]:
            v = str(v).strip()
            if v:
                allowed.append(v)

        seen = set()
        allowed = [x for x in allowed if not (x in seen or seen.add(x))]
        aliases[primary] = allowed

    return aliases


def resolve_match_pn(primary_pn: str, installed_by_pn: dict, aliases: dict):
    allowed = aliases.get(primary_pn, [primary_pn])
    for candidate in allowed:
        if candidate in installed_by_pn:
            return candidate
    return ""


def pick_description_from_export(rows_for_match: pd.DataFrame) -> str:
    if rows_for_match is None or rows_for_match.empty:
        return ""
    s = rows_for_match["Component Description"].astype(str).str.strip()
    s = s[s != ""]
    if s.empty:
        return ""
    return s.value_counts().idxmax()


def pick_description_from_combined(export_df: pd.DataFrame, primary_pn: str, aliases: dict) -> str:
    """
    Use the missing component PN to pull description from Combined as built (FULL export).
    Strategy:
      1) exact ProductNo match on primary + aliases
      2) base PN startswith match (handles rev differences)
    """
    prod = export_df["ProductNo"].astype(str).str.strip()
    desc = export_df["Component Description"].astype(str).str.strip()

    allowed = aliases.get(primary_pn, [primary_pn])

    def best(mask) -> str:
        if not mask.any():
            return ""
        s = desc[mask]
        s = s[s != ""]
        if s.empty:
            return ""
        return s.value_counts().idxmax()

    # exact matches first
    for cand in allowed:
        c = str(cand).strip()
        if not c:
            continue
        d = best(prod == c)
        if d:
            return d

    # base PN match next (e.g., 150405-01 -> 150405)
    for cand in allowed:
        b = str(cand).strip().split("-", 1)[0]
        if not b:
            continue
        d = best(prod.str.startswith(b))
        if d:
            return d

    return ""


def match_one_pn(installed_rows: pd.DataFrame, sn_lot_type: str):
    """
    Returns (qty_actual, ids_display_list)

    SN: qty = unique serial count, ids = unique serials
    LOT: qty = sum(Quantity) per lot, ids = ["LOT (xQ)", ...]
    """
    if installed_rows is None or installed_rows.empty:
        return 0, []

    if sn_lot_type == "SN":
        ids = [str(x).strip() for x in installed_rows["Serial #"].tolist() if str(x).strip()]
        uniq = sorted(set(ids))
        return len(uniq), uniq

    lot_mask = installed_rows["Lot #"].astype(str).str.strip() != ""
    lot_rows = installed_rows[lot_mask].copy()
    if lot_rows.empty:
        return 0, []

    lot_rows["Lot_clean"] = lot_rows["Lot #"].astype(str).str.strip()
    lot_rows["Qty_num"] = lot_rows["Quantity"].apply(to_float_or_zero)

    per_lot = lot_rows.groupby("Lot_clean")["Qty_num"].sum()

    qty_actual = int(per_lot.sum())
    display = [f"{lot} (x{int(q)})" for lot, q in per_lot.items()]
    return qty_actual, display


def get_run_parent_serial_from_anchor(df: pd.DataFrame, anchor_pn: str) -> str:
    pn_clean = df["ProductNo"].astype(str).str.strip()
    anchor_rows = df[pn_clean == anchor_pn]

    if anchor_rows.empty:
        raise ValueError(f"Anchor PN '{anchor_pn}' not found in export.")

    parents = anchor_rows["Parent Serial #"].astype(str).str.strip()
    parents = parents[parents != ""]

    if parents.empty:
        raise ValueError(f"Anchor PN '{anchor_pn}' has no Parent Serial # values.")

    return parents.value_counts().idxmax()


def write_bom_xlsx(output_path: str, export_df: pd.DataFrame, results: list[dict], run_parent_serial: str):
    wb = Workbook()

    # --- Sheet 1: Combined as built (FULL EXPORT) ---
    ws1 = wb.active
    ws1.title = "Combined as built"
    ws1.append(list(export_df.columns))
    for row in export_df.itertuples(index=False):
        ws1.append(list(row))

    # --- Sheet 2: 44807 ---
    ws2 = wb.create_sheet("44807")
    ws2.append([
        "PrimaryPartNo",
        "MatchedPartNo",
        "Component Description",
        "SN/LOT",
        "QTY Required",
        "QTY Actual",
        "IDs",
        "Status",
    ])
    for r in results:
        ws2.append([
            r["primary_pn"],
            r["matched_pn"],
            r["description"],
            r["snlot"],
            r["qty_required"],
            r["qty_actual"],
            r["ids_text"],
            r["status"],
        ])

    # --- Sheet 3: Vehicle intake (ONLY IDs that exist; NO missing lines; remove (xN)) ---
    ws3 = wb.create_sheet("Vehicle intake")
    ws3.append(["Parent Serial #", "Serial/Lot (from 44807)"])

    for r in results:
        if r["status"] == "NOT FOUND":
            continue

        ids_text = str(r.get("ids_text", "")).strip()
        if not ids_text:
            continue

        for line in ids_text.splitlines():
            clean = strip_lot_qty_suffix(line)
            if clean:
                ws3.append([run_parent_serial, clean])

    # --- Sheet 4: Missing Components (NOT FOUND + NOT SATISFIED) ---
    ws4 = wb.create_sheet("Missing Components")
    ws4.append([
        "Parent Serial #",
        "PartNo",
        "Description",
        "SN/LOT",
        "QTY Required",
        "QTY Actual",
        "Missing Qty",
        "Status",
    ])

    for r in results:
        if r["status"] == "SATISFIED":
            continue

        req = int(r.get("qty_required", 0))
        act = int(r.get("qty_actual", 0))
        missing_qty = max(0, req - act)

        ws4.append([
            run_parent_serial,
            r.get("primary_pn", ""),
            r.get("description", ""),
            r.get("snlot", ""),
            req,
            act,
            missing_qty,
            r.get("status", ""),
        ])

    wb.save(output_path)
    print("Saved BOM:", output_path)


if __name__ == "__main__":
    os.makedirs("outputs", exist_ok=True)

    reqs = read_requirements()
    print("Total requirements:", len(reqs))

    df = load_export()

    needed_cols = ["ProductNo", "Serial #", "Lot #", "Component Description", "Quantity", "Parent Serial #"]
    missing_cols = [c for c in needed_cols if c not in df.columns]
    if missing_cols:
        print("ERROR: export missing columns:", missing_cols)
        print("Available columns:", list(df.columns))
        raise SystemExit(1)

    ANCHOR_PN = "LBL-F5-01"
    run_parent_serial = get_run_parent_serial_from_anchor(df, ANCHOR_PN)
    print("Run Parent Serial (from anchor):", run_parent_serial)

    # Global matching (your working mode)
    df["ProductNo_clean"] = df["ProductNo"].astype(str).str.strip()
    installed_by_pn = {pn: g for pn, g in df.groupby("ProductNo_clean")}
    print("Unique ProductNo count:", len(installed_by_pn))

    aliases = load_aliases()
    print("Aliases loaded:", len(aliases))

    results = []
    for primary_pn, template_desc, qty_req, snlot in reqs:
        if qty_req == 0:
            continue

        matched_pn = resolve_match_pn(primary_pn, installed_by_pn, aliases)
        rows_for_match = installed_by_pn.get(matched_pn) if matched_pn else None

        qty_act, ids_display = match_one_pn(rows_for_match, snlot)

        export_desc = pick_description_from_export(rows_for_match)

        if matched_pn == "":
            # NOT FOUND: use the missing PN to pull description from Combined as built
            desc = pick_description_from_combined(df, primary_pn, aliases)
        else:
            # Found/not satisfied: prefer matched rows, fallback to template
            desc = export_desc if export_desc else template_desc

        if matched_pn == "":
            status = "NOT FOUND"
            qty_act = 0
            ids_display = []
        else:
            status = "SATISFIED" if qty_act >= qty_req else "NOT SATISFIED"

        results.append({
            "primary_pn": primary_pn,
            "matched_pn": matched_pn,
            "description": desc,
            "snlot": snlot,
            "qty_required": qty_req,
            "qty_actual": int(qty_act),
            "ids_text": "\n".join(ids_display),
            "status": status,
        })

    bom_path = f"outputs/{run_parent_serial}_BOM.xlsx"
    write_bom_xlsx(bom_path, df, results, run_parent_serial)

    print("Done.")
